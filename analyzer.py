# coding=utf-8

from typing import List
import os
import time
# import xlwt, xlrd
import openpyxl
from xlutils.copy import copy as xl_copy
import random
import re
import copy
import chardet

class Analyzer:
    def __init__(self, src: str, dst: str = "", sheet_name: str = "Sheet1",
                 encoding: str = None, greed_mode: bool = False):
        assert os.path.exists(src)
        self.__src = src
        self.__sheet_name = sheet_name
        self.__tab = None
        self.__title = []
        self.__title_map = {}
        self.__start_marker = ""
        self.__end_marker = ""
        self.__filter_map = {}
        self.__is_done = False
        self.__encoding = encoding if encoding else None
        self.__greed_mode = greed_mode
        self.raw_total_line = 0
        self.raw_read_line = 0
        self.output_write_line = 0

        if os.path.isfile(dst):
            self.__dst = dst
            self.__open_exist()
        else:
            if os.path.isdir(dst):
                self.__dst = dst.strip(os.sep).strip("/") + os.sep + \
                             os.path.splitext(os.path.split(src)[1])[0] + '.xlsx'
            else:
                self.__dst = os.path.split(src)[0] + os.sep + os.path.splitext(os.path.split(src)[1])[0] + '.xlsx'
            self.__output = openpyxl.Workbook()
            self.__tab = self.__output.active

    def __open_exist(self):
        dst = self.__dst
        sheet_name = self.__sheet_name
        if 1:
            rb = openpyxl.open(dst)
            self.__output = rb
            try:
                self.__tab = self.__output.get_sheet_by_name(sheet_name)
            except KeyError:
                self.__tab = self.__output.create_sheet(title=sheet_name)

    def set_title(self, title: str or List[str]):
        if isinstance(title, str):
            self.__title = title.split(",")
        elif isinstance(title, list):
            self.__title = title
        else:
            raise TypeError("not support title type:{0} of {1}".format(type(title), title))
        for i, item in enumerate(self.__title):
            self.__title_map[item] = i
        # 已有标题写入
        if self.__tab.max_row >= 2:
            return
        # 生成标题
        for i, item in enumerate(self.__title):
            self.__tab.cell(0 + 1, i + 1, item)
        self.__output.save(self.__dst)

    def set_start_end_marker(self, start_marker: str or re.Pattern, end_marker: str or re.Pattern):
        self.__start_marker = start_marker
        self.__end_marker = end_marker

    def add_capture_item(self, column_name: str, item_filter: str or re.Pattern,
                         index_in_filter: int = 0, skip_times: int = 0):
        self.__filter_map[(item_filter, index_in_filter, skip_times)] = column_name

    def _run(self, time_out: float = 60):
        assert self.__start_marker
        assert self.__end_marker
        # 自动识别文件编码
        my_encoding = self.__encoding
        if not my_encoding:
            with open(self.__src, 'rb') as f:
                my_encoding = chardet.detect(f.read(200))['encoding']

        print("use encoding = ", my_encoding, "\n")
        with open(self.__src, "r", encoding=my_encoding) as f:
            self.raw_total_line = f.readlines().__len__()
            print("{0} total lines is : {1}".format(self.__src, self.raw_total_line))
        is_in_session = False
        # row_index = 1
        # 从空行开始
        row_index = self.__tab.max_row
        search_map = {}
        with open(self.__src, "r", encoding=my_encoding) as f:
            st = time.time()
            while time.time() - st < time_out:
                line = f.readline()
                self.raw_read_line += 1
                if not line:
                    break
                be_trig = False  # 由查找开始标志到查找到跳变标志
                # if not is_in_session:
                if True:  # to check 任何时候查找到开始标志都开始新的匹配
                    if re.search(self.__start_marker, line):
                        if is_in_session:  # 未匹配到结束标志就开始
                            if self.__greed_mode:  # 贪婪模式则保存上一行
                                row_index += 1
                            else:  # 否则清空该行
                                for ii in range(self.__title_map.__len__()):
                                    self.__tab.cell(row_index + 1, ii + 1, "")
                        del search_map
                        search_map = copy.copy(self.__filter_map)
                        self.output_write_line += 1
                        is_in_session = True
                        be_trig = True
                if is_in_session:
                    for k, v in list(search_map.items()):
                        pattern, index, skip_times = k
                        sc_res = re.search(pattern, line)
                        if sc_res and sc_res.groups():
                            # print(sc_res.groups())
                            search_map.pop(k)
                            skip_times -= 1
                            if skip_times < 0 and sc_res.groups().__len__() > index:
                                #  assert sc_res.groups().__len__() > index
                                tab_index = self.__title_map[v]
                                wrt_item = sc_res.groups()[index]
                                try:
                                    self.__tab.cell(row_index + 1, tab_index + 1, float(wrt_item))
                                except ValueError:
                                    self.__tab.cell(row_index + 1, tab_index + 1, wrt_item)
                            else:
                                search_map[(pattern, index, skip_times)] = v

                    # 查找是否结束 (注 be_trig标志用于防止起始匹配和结束匹配相同时情况)
                    if re.search(self.__end_marker, line) and not be_trig:
                        is_in_session = False
                        row_index += 1

        self.__output.save(self.__dst)
        print("done ", self.__dst)
        self.__is_done = True

    def start(self):
        import threading

        def __t():
            try:
                self._run()
            except Exception as e:
                print(e)
                self.__is_done = True
        t = threading.Thread(target=__t)
        t.setDaemon(True)
        self.__is_done = False
        t.start()

    def is_done(self):
        return self.__is_done





"""
if __name__ == "__main__":
    a = Analyzer("test.txt", "output.xls")
    title = "起始时间,量程,测量类型,环境温度,质控P,质控Q,质控U2,3试剂进液时间," \
            "4试剂进液时间,硫酸温度1,硫酸温度2,显色后P,显色后Q,质控U1,测量值"
    a.set_title(title)
    a.set_start_end_marker(
        re.compile("([0-9\- :]*), \[(运行|维护)\] 启动-"),
        re.compile("([0-9\- :]*), \[(运行|维护)\] 结束-")
    )
    a.add_capture_item(
        "起始时间",
        re.compile("([0-9\- :]*), \[(运行|维护)\] 启动-"),
        0,
        0
    )
    a.start()
    while True:
        time.sleep(1)
    #pt = re.compile("([0-9\- :]*), \[(运行|维护)\] 启动-")
    #src = "2021-05-26 00:53:13, [维护] 启动-[0-250mg/L]标定K值,"
    #res = re.search(pt, src)
    #print(res)
    #print(res.groups())
"""